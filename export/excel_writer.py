"""Write SAP Data Mate staging workbook."""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="003366")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_ALT_FILL = PatternFill("solid", fgColor="EBF0FA")
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_sheet(ws, headers: list[str], rows: list[list]):
    """Apply consistent styling to a worksheet."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for row_idx, row in enumerate(rows, 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if fill:
                cell.fill = fill

    # Auto-width
    for col_idx, header in enumerate(headers, 1):
        max_len = max(
            len(str(header)),
            *[len(str(r[col_idx - 1])) for r in rows if col_idx - 1 < len(r)],
            1,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_excel(
    output: io.BytesIO,
    records: list[dict],
    rules_audit=None,
    include_trace: bool = True,
):
    """Write a multi-sheet Data Mate staging workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ── Sheet 1: Maintenance Plans ────────────────────────────────────────────
    ws_plans = wb.create_sheet("Maintenance Plans")
    plan_headers = [
        "Plan Name", "Plan Description", "Frequency", "Frequency Unit",
        "Item Description", "Is Regulatory", "Is Online", "Total Duration (hrs)",
    ]
    plan_rows = []
    seen_plan_items = set()
    for r in records:
        key = (r["Plan Name"], r["Item Description"])
        if key not in seen_plan_items:
            seen_plan_items.add(key)
            plan_rows.append([
                r["Plan Name"], r["Plan Description"],
                r["Frequency"], r["Frequency Unit"],
                r["Item Description"],
                "Yes" if r["Is Regulatory"] else "No",
                "Yes" if r["Is Online"] else "No",
                r["Total Duration (hrs)"],
            ])
    _style_sheet(ws_plans, plan_headers, plan_rows)

    # ── Sheet 2: Task Lists ───────────────────────────────────────────────────
    ws_tl = wb.create_sheet("Task Lists")
    tl_headers = ["Task List Name", "Plan Name", "Item Description"]
    tl_rows = []
    seen_tl = set()
    for r in records:
        key = r["Task List Name"]
        if key not in seen_tl:
            seen_tl.add(key)
            tl_rows.append([r["Task List Name"], r["Plan Name"], r["Item Description"]])
    _style_sheet(ws_tl, tl_headers, tl_rows)

    # ── Sheet 3: Operations ───────────────────────────────────────────────────
    ws_ops = wb.create_sheet("Operations")
    ops_headers = [
        "Task List Name", "Operation No", "Operation Description",
        "Resource Type", "Duration (hrs)", "Materials",
    ]
    ops_rows = [
        [
            r["Task List Name"], r["Operation No"], r["Operation Description"],
            r["Resource Type"], r["Duration (hrs)"], r["Materials"],
        ]
        for r in records
    ]
    _style_sheet(ws_ops, ops_headers, ops_rows)

    # ── Sheet 4: FMECA Traceability ───────────────────────────────────────────
    if include_trace:
        ws_trace = wb.create_sheet("FMECA Traceability")
        trace_headers = [
            "Plan Name", "Task List Name", "Operation No",
            "FLOC", "Failure Mode", "Criticality", "Source Task Type",
        ]
        trace_rows = [
            [
                r["Plan Name"], r["Task List Name"], r["Operation No"],
                r["FLOC"], r["Failure Mode"], r["Criticality"], r["Source Task Type"],
            ]
            for r in records
        ]
        _style_sheet(ws_trace, trace_headers, trace_rows)

    # ── Sheet 5: Rule Audit ───────────────────────────────────────────────────
    if rules_audit:
        ws_audit = wb.create_sheet("Rule Audit")
        audit_headers = ["Rule Type", "Description", "Parameter", "Value"]
        audit_rows = [
            [r["Rule Type"], r["Description"], r["Parameter"], r["Value"]]
            for r in rules_audit
        ]
        _style_sheet(ws_audit, audit_headers, audit_rows)

    wb.save(output)
